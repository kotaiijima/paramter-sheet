from flask import Flask, render_template, request
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
import datetime
import logging
import os
from tkinter import filedialog

LOGFILE_NAME = "DEBUG.log"

app = Flask(__name__)

app.logger.setLevel(logging.WARN)
log_handler = logging.FileHandler(LOGFILE_NAME)
log_handler.setLevel(logging.WARN)
app.logger.addHandler(log_handler)

#印刷範囲
print_area = 'A1:AU71'

@app.route("/", methods=["GET", "POST"])
def home():
    message = ""

    # ファイル名のフォーマット修正
    excel_file_name = 'パラメーターシート-%s.xlsx' % "パーキテック"

    # Excelファイルの作成と保存
    create_file = openpyxl.Workbook()
    create_file.save(excel_file_name)
    excel_file = openpyxl.load_workbook(excel_file_name)

    # シート名
    sheet_list = ['表紙', '本体設定', '電源設定', 'バックアップ設定']

    for sheet_name in sheet_list:
        excel_file.create_sheet(title=sheet_name)	

    excel_file.remove(excel_file['Sheet'])	
    excel_file.save(excel_file_name)

    # 書き込むExcelシートを選択
    sheet = excel_file[sheet_list[1]]
    #印刷範囲を指定
    sheet.print_area = print_area
    sheet.print_options.page_breaks = True
    sheet.page_setup.scale = 60
    # 書き込み始める行と列を定義
    start_row = 1
    start_col = 1
    end_row = 47
    end_col = 71
    #行と列の幅を調整
    row_height = 17.5
    col_width = 2.5
    #セルの枠線を非表示、１ページの枠にする
    sheet.sheet_view.showGridLines = False
    for y in range (start_col, end_col + 1):
        for x in range (start_row, end_row + 1):
            cell = sheet.cell(row=y, column=x)
            if y == 1:
                sheet.column_dimensions[cell.column_letter].width = col_width
            sheet.row_dimensions[y].height = row_height                

    excel_file.save(excel_file_name)

    #セルの塗りつぶし設定
    pf_blue1 = PatternFill(start_color="8EA9D8", end_color="8EA9D8", fill_type="solid")
    #塗りつぶし枠の範囲を設定
    frame1 = 38
    
    if request.method == "POST":
        print(request.form)
        if "submit_button" in request.form and request.form["submit_button"] == "clicked":
            item_nest_list = []
            print(len(request.form.getlist("items[0][]")))
            for i in range(len(request.form.getlist("items[0][]"))):  # 先頭要素の数からリストサイズ推測
                item_nest_list.append(request.form.getlist(f"items[{i}][]"))
            print(item_nest_list)
        # 書き込み
        #for y in range (1, frame1 + 1):
        #    for x, cell_value in enumerate(row):
        #        cell = sheet.cell(row=start_row + y, column=start_col + x, value=cell_value)	
        #        cell.fill = pf_blue1

    # ファイル保存ダイアログを表示（ファイル名も指定可能）

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Text files", "*.xlsx"), ("All files", "*.*")],
                title="保存先を指定"
            )

            if filepath:
                excel_file.save(filepath)
                message = print(f"ファイルを保存しました: {filepath}")

    os.remove(excel_file_name)
    
    return render_template("home.html", message=message)

if __name__ == '__main__':
    app.run(debug=True)