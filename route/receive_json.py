from flask import Flask, request, render_template, Blueprint, jsonify
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
import datetime
import logging
import os
from tkinter import filedialog
import tkinter
import xlwings 

app1 = Blueprint('receive_json', __name__, template_folder='templates')

#印刷範囲
print_area = 'A1:AU71'
# ファイル名のフォーマット修正
excel_file_name = "temp.xlsx"

@app1.route('/receive_json', methods=['POST'])
def receive_json():
    data = request.get_json()  # JSONデータを取得
    print("Received JSON:", data)  # コンソールに出力

    # Excelファイルの作成と保存
    create_file = openpyxl.Workbook()
    create_file.save(excel_file_name)
    excel_file = openpyxl.load_workbook(excel_file_name)

 
    excel_file.create_sheet(title=data["title"])	

    excel_file.remove(excel_file['Sheet'])	
    excel_file.save(excel_file_name)
    #フォントを指定
    font = Font(name="メイリオ")
    # 書き込むExcelシートを選択
    sheet = excel_file[data["title"]]
    # 余白
    sheet.page_margins.left = 0.8 / 2.54
    sheet.page_margins.right = 0.8 / 2.54
    sheet.page_margins.top = 0.9 / 2.54
    sheet.page_margins.bottom = 0.9 / 2.54
    sheet.page_margins.header = 0.8 / 2.54
    sheet.page_margins.footer = 0.8 / 2.54
    #印刷範囲を指定
    sheet.sheet_properties.pageSetUpPr.fitToPage = False
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.print_area = print_area
    sheet.print_options.page_breaks = True
    sheet.page_setup.scale = 60

    # 書き込み始める行と列を定義
    start_row = 1
    start_col = 1
    end_row = 47
    end_col = 71
    #行と列の幅を調整
    row_height = 17.5
    #col_width = 2.5
    col_width = 3.5
    #セルの枠線を非表示、１ページの枠にする
    sheet.sheet_view.showGridLines = False
    for y in range (start_col, end_col + 1):
        for x in range (start_row, end_row + 1):
            cell = sheet.cell(row=y, column=x)
            if y == 1:
                sheet.column_dimensions[cell.column_letter].width = col_width
            sheet.row_dimensions[y].height = row_height                

    excel_file.save(excel_file_name)

    #セルの塗りつぶし設定
    pf_blue1 = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    pf_blue2 = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    pf_blue3 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    #塗りつぶし枠の範囲を設定
    frame_range = 38
    #罫線の設定
    left_top_border = Border(
        top=Side(style="thin", color="000000"),
        left=Side(style="thin", color="000000")
        )
    
    right_top_border = Border(
        top=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000")
        )
    left_border = Border(left=Side(style="thin", color="000000"))
    right_border = Border(right=Side(style="thin", color="000000"))    
    top_border = Border(top=Side(style="thin", color="000000"))
    bottom_border = Border(bottom=Side(style="thin", color="000000"))

    # 書き込み始める行と列を定義
    write_row = 3
    write_col = 6
    write_col_middle = 26
    col_pos = 0
    #全体の行数を把握  
    total_row = 0

    #塗りつぶし範囲を把握
    column_array = []
    col_pos_minus = 0

    #現在の枠数
    now_frame = 0
    #各セルに取得した値を入力
    sheet["F2"] = data["title"]
    sheet["F2"].font = Font(name="メイリオ")

    #塗りつぶし範囲の調整    
    def check_col_pos_minus(y, count) :
        nonlocal col_pos_minus, now_frame
        col_pos_minus = 0
        for i in range(count): 
            if y >= column_array[i] : col_pos_minus += 1
        if count == 3: print(f"y = {y}, column_array={column_array[i]}, col_pos_minus:{col_pos_minus}")
    #ヘッダーの処理
    def header_func(frame):
        nonlocal total_row, col_pos, col_pos_minus
        check_col_pos_minus(total_row, col_pos)
        if col_pos == 3: print(f"total_row:{total_row} pos:{col_pos}, minus:{col_pos_minus} total={col_pos + write_col - col_pos_minus}  name:{frame}")
        cell = sheet.cell(row=write_row + total_row, column=write_col + col_pos - col_pos_minus, value=frame)
        cell.font = Font(bold=True, name="メイリオ")
        #枠の塗りつぶし
        for x in range(frame_range):
            cell = sheet.cell(row=write_row + total_row, column=(write_col + x))
            match (col_pos - col_pos_minus):
                case 0:
                    cell.fill = pf_blue1
                case 1:
                    if (x) == 0:
                        cell.fill = pf_blue1
                    else:
                        cell.fill = pf_blue2
                case _:
                    if (x) == 0:
                        cell.fill = pf_blue1
                    elif (x) == 1:
                        cell.fill = pf_blue2
                    else:
                        cell.fill = pf_blue3
        col_pos += 1
        total_row += 1

    for  frame in list(data.keys()):
        print(frame)
        if frame == "title": continue
        print(data[frame])
        if not data[frame] : header_func(frame=frame) 
        else:
            for index, item in enumerate(data[frame]):
                if index == 0: 
                    column_array.append(int(item["column"]) + total_row)
                    header_func(frame=frame)
                for key, value in item.items():
                    if key == "column": continue
                    check_col_pos_minus(total_row, col_pos)
                    cell = sheet.cell(row=write_row + total_row, column=write_col + col_pos - col_pos_minus, value=key)
                    cell.font = Font(name="メイリオ")
                    cell = sheet.cell(row=write_row + total_row, column=write_col_middle, value=value)
                    cell.font = Font(name="メイリオ")
                    #枠の塗りつぶし
                    for x in range(col_pos - col_pos_minus):
                        #if (write_col + x - col_pos_minus) < 6 : continue
                        cell = sheet.cell(row=write_row + total_row, column=write_col + x)
                        match (x):
                            case 0:
                                cell.fill = pf_blue1
                            case 1:
                                cell.fill = pf_blue2
                            case _:
                                cell.fill = pf_blue3
                    total_row += 1
    
    # 枠線を追加
    fill_count = 0
    for y in range(total_row + 1):
        cell = sheet.cell(row=write_row + y, column=10)
        #セルが塗りつぶされている（ヘッダー）か判別
        is_filled = (
            cell.fill.patternType == "solid"
            and cell.fill.fgColor.rgb not in (None, "00000000")
        )

        check_col_pos_minus(y, fill_count)
        #print(f"y={y}, fillcount={fill_count}, col_pos_minus={col_pos_minus}")
        for x in range(frame_range):
            cell = sheet.cell(row=write_row + y, column=write_col + x)
            if y != total_row:
                if (x - (fill_count - col_pos_minus)) < 0:
                    cell.border = left_border
                elif (x - (fill_count - col_pos_minus)) == 0:
                    cell.border = left_top_border
                elif x == (frame_range - 1):
                    cell.border = right_top_border
                else:
                    cell.border = top_border
            else:
                cell.border = top_border
        if is_filled:
            fill_count += 1
        elif y != total_row:
            cell = sheet.cell(row=write_row + y, column=write_col_middle)
            cell.border = left_top_border
    excel_file.save(excel_file_name)

    # ファイル保存ダイアログを表示（ファイル名も指定可能）
    root = tkinter.Tk()
    # topmost指定(最前面)
    root.attributes('-topmost', True)
    root.withdraw()
    root.lift()
    filepath = filedialog.asksaveasfilename(
        parent=root,
        defaultextension=".xlsx",
        filetypes=[("Text files", "*.xlsx"), ("All files", "*.*")],
        title="保存先を指定"
    )
    root.focus_force()

    if filepath:
        if os.path.exists(filepath):
            with xlwings.App(visible=False) as app:
                src_wb = xlwings.Book(r'temp.xlsx')
                dst_wb = xlwings.Book(filepath)
                src_wb.sheets[0].copy(after=dst_wb.sheets[0])
                dst_wb.save()
                src_wb.close()
                dst_wb.close()
        else:
            excel_file.save(filepath)
            message = print(f"ファイルを保存しました: {filepath}")
    else:
        return "", 204
    
    #tempファイルを削除
    os.remove(excel_file_name)

#    return render_template("home.html", message=message)
    return jsonify({"message": "JSON received", "data": data})  # レスポンスを返す

if __name__ == '__main__':
    app1.run(debug=True)